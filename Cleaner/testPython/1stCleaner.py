import os
import re
import pandas as pd
from openpyxl import load_workbook

# Input and output
input_folder = "ExcelFolders"
output_file = os.path.join("cleanExcel", "cleanedBook.xlsx")
log_file = "log.txt"

# Make sure output folder exists
os.makedirs("cleanExcel", exist_ok=True)

# Load log of already processed files
if os.path.exists(log_file):
    with open(log_file, "r") as f:
        processed_files = set(f.read().splitlines())
else:
    processed_files = set()

# Collect new cleaned DataFrames
new_dfs = []

# Loop through Excel files and skip ones already processed
for filename in os.listdir(input_folder):
    if filename.endswith(".xlsx") and filename not in processed_files:
        filepath = os.path.join(input_folder, filename)

        # 1. Load workbook and unmerge cells
        wb = load_workbook(filepath)
        ws = wb.active
        for merged_range in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged_range))
        wb.save(filepath)

        # 2. Read Excel into DataFrame
        df = pd.read_excel(filepath, header=None)
        df = df.dropna(how="all")  # remove empty rows

        if df.shape[1] > 0:
            df = df.drop(df.columns[0], axis=1)  # drop first col
        df.insert(0, "EmptyCol1", "")
        df.insert(0, "EmptyCol2", "")

        new_dfs.append(df)

        # Add file to log
        with open(log_file, "a") as f:
            f.write(filename + "\n")

# If no new files, exit
if not new_dfs:
    print("⚠ No new Excel files to process.")
    exit()

# Load existing master if exists
if os.path.exists(output_file):
    master_df = pd.read_excel(output_file)
else:
    master_df = pd.DataFrame()

# Combine old + new data
final_df = pd.concat([master_df] + new_dfs, ignore_index=True)

# Keep only 36 columns
final_df = final_df.iloc[:, :36]

# Apply final cleaning steps
if final_df.shape[1] > 35:
    final_df = final_df.drop(final_df.columns[35], axis=1)

final_df.columns = [
    "Month_year", "Consultation_Type", "Case",
    "Under 1 Male", "Under 1 Female",
    "1-4 Male", "1-4 Female",
    "5-9 Male", "5-9 Female",
    "10-14 Male", "10-14 Female",
    "15-18 Male", "15-18 Female",
    "19-24 Male", "19-24 Female",
    "25-29 Male", "25-29 Female",
    "30-34 Male", "30-34 Female",
    "35-39 Male", "35-39 Female",
    "40-44 Male", "40-44 Female",
    "45-49 Male", "45-49 Female",
    "50-54 Male", "50-54 Female",
    "55-59 Male", "55-59 Female",
    "60-64 Male", "60-64 Female",
    "65-69 Male", "65-69 Female",
    "70 Over Male", "70 Over Female"
]

# Fill NaN with 0
final_df = final_df.fillna(0)

# Save back to master
final_df.to_excel(output_file, index=False)

print(f"✅ Master file updated with {len(new_dfs)} new files → {output_file}")
